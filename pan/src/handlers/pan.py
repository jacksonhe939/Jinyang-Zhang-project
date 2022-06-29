import re
import os
import json
import time
import datetime

from openpyxl import load_workbook

from utils import req
from config import settings


class PanHandler(object):

    def __init__(self, conn):
        # client version socket connect
        self.conn = conn
        self.username = None

    @property
    def home_path(self):
        return os.path.join(settings.USER_FOLDER_PATH, self.username)

    def send_json_data(self, **kwargs):
        # kwargs={"status": False, 'error': "user exist"}
        req.send_data(self.conn, json.dumps(kwargs))

    def recv_save_file(self, target_file_path):
        req.recv_save_file(self.conn, target_file_path)

    def send_file_by_seek(self, file_size, file_path, seek=0):
        req.send_file_by_seek(self.conn, file_size, file_path, seek)

    def login(self, username, password):
        """ user login，read excel file，user login """
        wb = load_workbook(settings.DB_FILE_PATH)
        sheet = wb.worksheets[0]

        success = False
        for row in sheet.iter_rows(2):
            if username == row[0].value and password == row[1].value:
                success = True
                break

        if success:
            # req.send_data(self.conn, json.dumps({"status": True, 'data': "login successfully"}))
            self.send_json_data(status=True, data="login successfully")
            self.username = username
        else:
            # req.send_data(self.conn, json.dumps({"status": False, 'error': "login failed"}))
            self.send_json_data(status=False, error="login failed")

    def register(self, username, password):
        """ user register， user and password write into excel（if exist not going to register ） """

        wb = load_workbook(settings.DB_FILE_PATH)
        sheet = wb.worksheets[0]

        # check user exist or not
        exists = False
        for row in sheet.iter_rows(2):
            if username == row[0].value:
                exists = True
                break
        if exists:

            # self.conn.sendall(b"....") -> sticky package
            # req.send_data(self.conn, json.dumps({"status": False, 'error': "user exist"}))
            self.send_json_data(status=False, error="user exist")
            return

        # registered user write into Excel
        max_row = sheet.max_row
        data_list = [username, password, datetime.datetime.now().strftime("%Y-%m-%d")]
        for i, item in enumerate(data_list, 1):
            cell = sheet.cell(max_row + 1, i)
            cell.value = item
        wb.save(settings.DB_FILE_PATH)

        # create user folder
        user_folder = os.path.join(settings.USER_FOLDER_PATH, username)
        os.makedirs(user_folder)

        # reply message
        # req.send_data(self.conn, json.dumps({"status": True, 'data': "successfully registered"}))
        self.send_json_data(status=True, data="successfully registered")

    def ls(self, folder_path=None):
        """ check user's files
            1. folder_path=None，check root file
            2. folder_path not none check user's file/folder_path
        """
        if not self.username:
            self.send_json_data(status=False, error="login then check it")
            return

        if not folder_path:
            # root file： files + username
            data = "\n".join(os.listdir(self.home_path))
            self.send_json_data(status=True, data=data)
            return

        target_folder = os.path.join(self.home_path, folder_path)

        if not os.path.exists(target_folder):
            self.send_json_data(status=False, error="file path not exist")
            return
        if not os.path.isdir(target_folder):
            self.send_json_data(status=False, error="folder not exist")
            return

        data = "\n".join(os.listdir(target_folder))
        self.send_json_data(status=True, data=data)

    def upload(self, file_path):
        """ upload file，override"""
        # 用户未登录
        if not self.username:
            self.send_json_data(status=False, error="login then check it")
            return

        target_file_path = os.path.join(self.home_path, file_path)
        folder = os.path.dirname(target_file_path)
        if not os.path.exists(folder):
            os.makedirs(folder)

        self.send_json_data(status=True, data="start upload")

        # receive file
        self.recv_save_file(target_file_path)

    def download(self, file_path, seek=0):
        """ download file，
            seek=None，start from beginning；
            seek=1000，from 1000 bits download
        """
        # user not login
        if not self.username:
            # req.send_data(self.conn, json.dumps({"status": False, "error": "login first then upload"}))
            self.send_json_data(status=False, error="login first then upload")
            return

        # file not exist
        target_file_path = os.path.join(self.home_path, file_path)
        if not os.path.exists(target_file_path):
            # req.send_data(self.conn, json.dumps({"status": False, "error": "文件{}不存在".format(file_path)}))
            self.send_json_data(status=False, error="file {} not exist".format(file_path))
            return

        # get file_size and return
        # req.send_data(self.conn, json.dumps({"status": True, "data": "start download"}))
        self.send_json_data(status=True, data="start download")

        # send file
        seek = int(seek)
        total_size = os.stat(target_file_path).st_size
        req.send_file_by_seek(self.conn, total_size - seek, target_file_path, seek)

    def execute(self):
        """
        client send requests，trigger this method。
        :return: False，shut down connection；True，continue requests
        """
        conn = self.conn

        # login/register/check file/upload/download

        # 1.get package
        cmd = req.recv_data(conn).decode('utf-8')
        if cmd.upper() == "Q":
            print("client exit")
            return False

        method_map = {
            "login": self.login,
            "register": self.register,
            "ls": self.ls,
            "upload": self.upload,
            "download": self.download,
        }

        # "register  root   666"
        # "login  alex   123"
        # "ls"      [ls,]
        # "ls xxx"  [ls,xxx]
        # upload v2.py
        cmd, *args = re.split(r"\s+", cmd)  # [register,root,123] /  [login,root,123]
        method = method_map[cmd]

        method(*args)

        return True
